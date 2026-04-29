#!/usr/bin/env python3
"""
Build XLSX fixtures.

Each fixture targets specific behavior. We document what extract-text does
vs. what we want our impl to do — there are several places we choose to
improve over its baseline (notably date formatting).
"""
from pathlib import Path
from datetime import datetime, date, time
from openpyxl import Workbook
import zipfile

OUT = Path(__file__).resolve().parent.parent / "fixtures" / "xlsx"
OUT.mkdir(parents=True, exist_ok=True)


# ---------- 01: simple data, single sheet ----------
def build_01_basic():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score", "Note"])
    ws.append(["Alice", 95, "first"])
    ws.append(["Bob", 80, "second"])
    ws.append(["Carol", 70, "third"])
    wb.save(OUT / "01_basic.xlsx")


# ---------- 02: multiple sheets, including empty one ----------
def build_02_multi_sheets():
    wb = Workbook()
    ws = wb.active
    ws.title = "First"
    ws.append(["a", "b"])
    ws.append([1, 2])
    ws2 = wb.create_sheet("Second")
    ws2.append(["x", "y"])
    ws2.append([3, 4])
    wb.create_sheet("Empty")
    wb.save(OUT / "02_multi_sheets.xlsx")


# ---------- 03: types — number, date, bool, formula, error ----------
def build_03_types():
    wb = Workbook()
    ws = wb.active
    ws.title = "Types"
    ws.append(["str", 12345.6789, True, datetime(2025, 1, 15, 14, 30, 0)])
    ws.append([1000000, 0.0000001, False, date(2025, 1, 16)])
    ws.append(["with newline\nhere", "with tab\there", "with pipe|here", time(14, 30)])
    wb.save(OUT / "03_types.xlsx")


# ---------- 04: sparse rows, merged cells, out-of-bounds ----------
def build_04_sparse_merged():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "header"
    ws["E1"] = "far"          # gap of 3 columns
    ws["A3"] = "row 3"
    ws["A6"] = "merged"
    ws.merge_cells("A6:C6")
    wb.save(OUT / "04_sparse_merged.xlsx")


# ---------- 05: formulas with cached values ----------
def build_05_formulas():
    """Hand-write XML to control cached <v> values precisely."""
    sheet = """<?xml version="1.0"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetData>
<row r="1"><c r="A1" t="s"><v>0</v></c><c r="B1"><v>10</v></c></row>
<row r="2"><c r="A2" t="str"><f>UPPER(A1)</f><v>HELLO</v></c><c r="B2"><f>B1*2</f><v>20</v></c></row>
<row r="3"><c r="A3" t="e"><v>#DIV/0!</v></c></row>
</sheetData></worksheet>"""
    sst = """<?xml version="1.0"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>hello</t></si></sst>"""
    write_minimal_xlsx(OUT / "05_formulas.xlsx", sheet, sst, sheet_name="Formulas")


# ---------- 06: empty workbook ----------
def build_06_empty():
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    wb.save(OUT / "06_empty.xlsx")


# ---------- 07: special characters in cell values (TSV/markdown safety) ----------
def build_07_special_chars():
    wb = Workbook()
    ws = wb.active
    ws.title = "Special"
    ws.append(["pipe |", "tab\there", "newline\nhere"])
    ws.append(["smart \u201cquote\u201d", "em — dash", "中文 mixed"])
    wb.save(OUT / "07_special_chars.xlsx")


# ---------- 08: shared strings deduplication ----------
def build_08_shared_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Shared"
    repeat = "repeated"
    for _ in range(5):
        ws.append([repeat, "unique" + str(_)])
    wb.save(OUT / "08_shared_strings.xlsx")


# ============================================================
def write_minimal_xlsx(path, sheet_xml, sst_xml=None, sheet_name="Sheet1"):
    content_types = """<?xml version="1.0"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>"""
    if sst_xml:
        content_types += '\n<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
    content_types += "\n</Types>"

    root_rels = """<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>"""
    workbook = f"""<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="{sheet_name}" sheetId="1" r:id="rIdS1"/></sheets></workbook>"""
    wb_rels_parts = ['<Relationship Id="rIdS1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>']
    if sst_xml:
        wb_rels_parts.append('<Relationship Id="rIdSST" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>')
    wb_rels = '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">' + "".join(wb_rels_parts) + "</Relationships>"

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", root_rels)
        z.writestr("xl/workbook.xml", workbook)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        if sst_xml:
            z.writestr("xl/sharedStrings.xml", sst_xml)


if __name__ == "__main__":
    for name, fn in list(globals().items()):
        if name.startswith("build_") and callable(fn):
            print(f"Building {name}...")
            fn()
